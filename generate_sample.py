import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import date

wb = openpyxl.Workbook()

# ── Colores corporativos Isopan ──
MARRON = "8B4513"
DORADO = "D4A017"
CREMA  = "FFF8F0"
GRIS   = "F2F2F2"

def header_style(cell, bg=MARRON):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def subheader_style(cell):
    cell.fill = PatternFill("solid", fgColor=DORADO)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")

thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# ── Hoja 1: Centros de Producción ──────────────────────────────
ws1 = wb.active
ws1.title = "Centros_Produccion"
ws1.column_dimensions["A"].width = 25
ws1.column_dimensions["B"].width = 20
ws1.column_dimensions["C"].width = 18
ws1.column_dimensions["D"].width = 18
ws1.column_dimensions["E"].width = 20

headers = ["Centro de Producción", "Responsable", "Turno", "Estado", "Capacidad (unid/día)"]
for c, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=c, value=h)
    header_style(cell)
    cell.border = thin

rows = [
    ["CP-01 Panadería Principal", "Carlos Pérez", "Mañana", "Activo", 2500],
    ["CP-02 Repostería", "María López", "Tarde", "Activo", 1800],
    ["CP-03 Pan Especial", "José García", "Mañana", "Activo", 1200],
    ["CP-04 Tortillería", "Ana Martínez", "Completo", "Activo", 3000],
    ["CP-05 Pastelería", "Luis Reyes", "Mañana", "Mantenimiento", 900],
]
for r, row in enumerate(rows, 2):
    for c, val in enumerate(row, 1):
        cell = ws1.cell(row=r, column=c, value=val)
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center")
        cell.fill = PatternFill("solid", fgColor=CREMA if r % 2 == 0 else "FFFFFF")
        cell.border = thin
ws1.row_dimensions[1].height = 35

# ── Hoja 2: Producción ─────────────────────────────────────────
ws2 = wb.create_sheet("Produccion")
for col, w in zip("ABCDEFGHIJ", [25, 20, 15, 15, 15, 15, 15, 12, 12, 18]):
    ws2.column_dimensions[col].width = w

headers2 = [
    "Producto", "Centro", "Planificado (u)",
    "Elaborado (u)", "Rendimiento (%)", "Merma (u)",
    "Merma (%)", "No Elaborado (u)", "Fecha", "Turno"
]
for c, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    header_style(cell)
    cell.border = thin
ws2.row_dimensions[1].height = 40

data = [
    ["Pan Francés 50g",     "CP-01", 1000, 980,  98.0, 12,  1.2,  20, str(date.today()), "Mañana"],
    ["Pan Dulce Surtido",   "CP-02",  800, 765,  95.6, 25,  3.1,  35, str(date.today()), "Tarde"],
    ["Croissant Mantequilla","CP-02", 500, 490,  98.0,  8,  1.6,  10, str(date.today()), "Mañana"],
    ["Tortilla de Maíz",    "CP-04", 2000,1950,  97.5, 40,  2.0,  50, str(date.today()), "Completo"],
    ["Pan de Molde",        "CP-01",  600, 580,  96.7, 15,  2.5,  20, str(date.today()), "Mañana"],
    ["Bolillo",             "CP-03",  900, 870,  96.7, 22,  2.4,  30, str(date.today()), "Tarde"],
    ["Pastel de Tres Leches","CP-05",  200, 185,  92.5, 10,  5.0,  15, str(date.today()), "Mañana"],
    ["Pan Integral",        "CP-01",  400, 395,  98.8,  4,  1.0,   5, str(date.today()), "Tarde"],
]
for r, row in enumerate(data, 2):
    for c, val in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center")
        cell.fill = PatternFill("solid", fgColor=CREMA if r % 2 == 0 else "FFFFFF")
        cell.border = thin

# ── Hoja 3: Materiales Pendientes ─────────────────────────────
ws3 = wb.create_sheet("Materiales_Pendientes")
for col, w in zip("ABCDEFG", [25, 20, 15, 15, 15, 20, 20]):
    ws3.column_dimensions[col].width = w

headers3 = ["Material", "Centro Solicitante", "Cantidad Solicitada",
            "Unidad", "Stock Actual", "Estado", "Fecha Solicitud"]
for c, h in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=c, value=h)
    header_style(cell)
    cell.border = thin
ws3.row_dimensions[1].height = 35

mat_data = [
    ["Harina de Trigo",   "CP-01", 500, "Quintales", 120, "URGENTE",  str(date.today())],
    ["Azúcar Blanca",     "CP-02", 200, "Quintales",  80, "Pendiente",str(date.today())],
    ["Mantequilla",       "CP-02", 100, "Cajas",      30, "Pendiente",str(date.today())],
    ["Levadura Seca",     "CP-01",  50, "Kg",         10, "URGENTE",  str(date.today())],
    ["Sal de Mesa",       "CP-03",  80, "Quintales",  50, "Normal",   str(date.today())],
    ["Masa de Maíz",      "CP-04", 300, "Quintales",  90, "Pendiente",str(date.today())],
    ["Esencia de Vainilla","CP-05",  20, "Litros",      5, "URGENTE",  str(date.today())],
]
for r, row in enumerate(mat_data, 2):
    for c, val in enumerate(row, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center")
        cell.fill = PatternFill("solid", fgColor=CREMA if r % 2 == 0 else "FFFFFF")
        cell.border = thin

today = date.today().strftime("%Y-%m-%d")
path = f"/home/claude/isopan-zona7/data/daily/produccion_{today}.xlsx"
wb.save(path)
print(f"Saved: {path}")
