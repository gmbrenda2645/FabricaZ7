import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import date, timedelta
import random

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Datos"

headers = ["Fecha", "Tienda", "Codigo", "Producto", "Devolucion", "Mala produccion", "Total", "Razon"]
ws.append(headers)

tiendas = ["FUTECA SAN CRISTOBAL", "FUTECA ZONA 7", "FUTECA MIXCO", "FUTECA VILLA NUEVA"]
productos = [
    ("220", "FRANCES PEQUENO 18"),
    ("221", "FRANCES GRANDE 18"),
    ("226", "BOLILLO 18"),
    ("234", "BAGUETTE PEQUENA 18"),
    ("237", "BAGUETTE GIGANTE 18"),
    ("240", "PAN DE AJO 18"),
    ("243", "PAN DE AGUA 2 ONZ"),
    ("269", "VOLOVAN DE POLLO 18"),
    ("270", "VOLOVAN DE CARNE 18"),
    ("403", "MILHOJAS 06"),
    ("406", "EMPANADA DE LECHE"),
    ("613", "BOQUITAS (BOLSA)"),
    ("615", "PAN SIN GLUTEN 09"),
    ("S150", "DONA UNIDAD"),
    ("U005", "PIZZA ROLLS UNIDAD"),
]
razones = ["Devolucion", "Cadena de frio", "Apachado", "Horneo", ""]

start = date(2026, 6, 1)
rows = []
for d in range(32):
    fecha = start + timedelta(days=d)
    for tienda in tiendas:
        n_items = random.randint(4, 9)
        for cod, prod in random.sample(productos, n_items):
            dev = random.choice([0, 0, 1, 1, 2, 3, 5, 8])
            mp = random.choice([0, 0, 0, 1, 1, 2, 4])
            total = dev + mp
            if total == 0:
                continue
            razon = random.choice(razones) if total > 0 else ""
            if not razon and total > 0:
                razon = "Devolucion"
            rows.append([fecha, tienda, cod, prod, dev, mp, total, razon])

for r in rows:
    ws.append(r)

# Format date column
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    for cell in row:
        cell.number_format = "DD/MM/YYYY"

# Create Excel Table
last_row = ws.max_row
last_col_letter = "H"
table_ref = f"A1:{last_col_letter}{last_row}"
tab = Table(displayName="TablaDevoluciones", ref=table_ref)
style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
tab.tableStyleInfo = style
ws.add_table(tab)

for col_cells in ws.columns:
    max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
    ws.column_dimensions[col_cells[0].column_letter].width = max_len + 3

wb.save("/home/claude/isopan-devoluciones/data/reporte_devolucion_pm.xlsx")
print(f"Generado con {len(rows)} filas")
